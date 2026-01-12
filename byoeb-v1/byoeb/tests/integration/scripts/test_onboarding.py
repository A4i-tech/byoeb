import sys

from openpyxl import Workbook, load_workbook
import pytest

# Skip entire test module if onboarding can't be imported (e.g., Azure Monitor dependency issue)
onboarding = pytest.importorskip("byoeb.scripts.onboarding", reason="onboarding module import failed (likely Azure Monitor dependency issue)")


DUMMY_PHONE_NUMBERS = ["9990000000001", "9990000000002"]


@pytest.fixture
def onboarding_excel(tmp_path) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "user_id",
        "user_name",
        "phone_number_id",
        "user_location",
        "user_type",
        "test_user",
        "created_timestamp",
        "user_language",
    ])
    sheet.append([
        "user-1",
        "First User",
        int(DUMMY_PHONE_NUMBERS[0]),
        "{'district': 'Test District 1'}",
        "asha",
        True,
        1_700_000_000,
        "en",
    ])
    sheet.append([
        "user-2",
        "Second User",
        int(DUMMY_PHONE_NUMBERS[1]),
        "{'district': 'Test District 2'}",
        "anm",
        False,
        1_700_000_100,
        "hi",
    ])

    file_path = tmp_path / "onboarding.xlsx"
    workbook.save(file_path)
    return str(file_path)


@pytest.fixture
def cleanup_dummy_users(envs, auth_session):
    auth_session.delete(f"{envs.base_url}/delete_users", json=DUMMY_PHONE_NUMBERS)
    yield DUMMY_PHONE_NUMBERS
    auth_session.delete(f"{envs.base_url}/delete_users", json=DUMMY_PHONE_NUMBERS)


def test_onboarding_registers_and_exports_users(monkeypatch, onboarding_excel, cleanup_dummy_users, tmp_path, envs, auth_session):
    output_sheet = tmp_path / "output.xlsx"
    monkeypatch.setattr(sys, "argv", ["onboarding.py", "--file", onboarding_excel, "--url", envs.base_url, "--sheet", str(output_sheet)])
    onboarding.main(session=auth_session)

    get_response = auth_session.post(f"{envs.base_url}/get_users", json=cleanup_dummy_users)
    get_response.raise_for_status()
    users = get_response.json()

    assert len(users) == len(cleanup_dummy_users)
    returned_numbers = {str(user.get("phone_number_id")) for user in users}
    assert returned_numbers == set(cleanup_dummy_users)
    assert all(
        isinstance(user.get("user_location"), dict) and user["user_location"].get("district")
        for user in users
    )

    assert output_sheet.exists()
    sheet = load_workbook(output_sheet).active
    rows = list(sheet.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    exported_rows = [dict(zip(header, row)) for row in data_rows]
    exported_numbers = {str(row.get("phone")) for row in exported_rows}
    assert exported_numbers == set(cleanup_dummy_users)
    assert all("district" in str(row.get("location", "")).lower() for row in exported_rows)


def test_onboarding_updates_users(monkeypatch, onboarding_excel, cleanup_dummy_users, envs, auth_session):
    monkeypatch.setattr(sys, "argv", ["onboarding.py", "--file", onboarding_excel, "--url", envs.base_url, "--update"])
    onboarding.main(session=auth_session)

    get_response = auth_session.post(f"{envs.base_url}/get_users", json=cleanup_dummy_users)
    get_response.raise_for_status()
    users = get_response.json()
    assert len(users) == len(cleanup_dummy_users)
